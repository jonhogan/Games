import openpyxl
import os

# Define the file path for excel file
fileName = input("Enter the file name with extension (example.xlsx): ")
current_directory = os.path.dirname(os.path.realpath(__file__))
excel_file_path = os.path.join(current_directory, fileName)

# Load excel file
wb = openpyxl.load_workbook(excel_file_path)
# define the sheet name
sheet = wb[input("Enter the sheet name: ")]
sheet['G1'].value = "Last Name"
sheet['H1'].value = "First Name"

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 6)
    name = cell.value.split(", ")
    sheet.cell(row, 7).value = name[0]
    sheet.cell(row, 8).value = name[1]

wb.save(excel_file_path)